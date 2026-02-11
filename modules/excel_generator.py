"""
SonIA Core - Excel Report Generator Module
Generates per-client and consolidated Excel tracking reports.
Replicates the format of SonIA Tracker's SonIA_Tracking_Results.xlsx
"""

import logging
import os
import tempfile
from datetime import datetime, timezone, timedelta, date
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COT = timezone(timedelta(hours=-5))

HEADERS = [
    "Nombre Cliente",
    "FEDEX Tracking",
    "SonIA status",
    "FedEx status",
    "Label Creation Date",
    "Shipping Date",
    "Days After Shipment",
    "Working Days After Shipment",
    "Days After Label Creation",
    "Destination City/State/Country",
    "Historial",
    "SonIA Recomendacion",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
DELIVERED_FILL = PatternFill("solid", fgColor="E2EFDA")
ALERT_FILL = PatternFill("solid", fgColor="FCE4EC")

COL_WIDTHS = [22, 18, 14, 20, 18, 16, 24, 28, 24, 30, 60, 40]

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


class ExcelReportGenerator:
    """Generates Excel tracking reports matching SonIA Tracker format."""

    def __init__(self, output_dir: str = "/tmp/sonia_reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_consolidated_report(
        self, all_shipments: Dict[str, List[Dict]], date_str: str = ""
    ) -> Optional[str]:
        """
        Generate consolidated Excel with all tenants' shipments.

        Args:
            all_shipments: {tenant_name: [shipment_dicts]}
            date_str: date string for filename

        Returns:
            File path of generated Excel, or None on error
        """
        if not date_str:
            date_str = datetime.now(COT).strftime("%Y-%m-%d")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Consolidado"

            self._write_headers(ws)

            row = 2
            for tenant_name in sorted(all_shipments.keys()):
                shipments = all_shipments[tenant_name]
                for s in shipments:
                    self._write_shipment_row(ws, row, tenant_name, s)
                    row += 1

            self._auto_filter(ws, row - 1)

            filepath = os.path.join(
                self.output_dir, f"SonIA_Tracking_Consolidado_{date_str}.xlsx"
            )
            wb.save(filepath)
            logger.info(f"Consolidated report saved: {filepath} ({row - 2} shipments)")
            return filepath

        except Exception as e:
            logger.error(f"Error generating consolidated report: {e}")
            return None

    def generate_tenant_report(
        self, tenant_name: str, shipments: List[Dict], date_str: str = ""
    ) -> Optional[str]:
        """
        Generate filtered Excel for a single tenant.

        Args:
            tenant_name: Client/tenant name
            shipments: List of shipment dicts
            date_str: date string for filename

        Returns:
            File path of generated Excel, or None on error
        """
        if not date_str:
            date_str = datetime.now(COT).strftime("%Y-%m-%d")

        if not shipments:
            logger.warning(f"No shipments to report for {tenant_name}")
            return None

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = tenant_name[:31]  # Excel sheet name max 31 chars

            self._write_headers(ws)

            for i, s in enumerate(shipments, start=2):
                self._write_shipment_row(ws, i, tenant_name, s)

            self._auto_filter(ws, len(shipments) + 1)

            safe_name = "".join(
                c if c.isalnum() or c in "_ -" else "_" for c in tenant_name
            )
            filepath = os.path.join(
                self.output_dir, f"SonIA_Tracking_{safe_name}_{date_str}.xlsx"
            )
            wb.save(filepath)
            logger.info(
                f"Tenant report saved: {filepath} ({len(shipments)} shipments)"
            )
            return filepath

        except Exception as e:
            logger.error(f"Error generating report for {tenant_name}: {e}")
            return None

    def _write_headers(self, ws):
        """Write header row with formatting."""
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def _write_shipment_row(self, ws, row: int, tenant_name: str, s: Dict):
        """Write a single shipment row."""
        now = datetime.now(COT).date()

        sonia_status = s.get("sonia_status", "unknown")
        fedex_status = s.get("fedex_status", "")
        label_date = self._parse_date(s.get("label_creation_date"))
        ship_date = self._parse_date(s.get("ship_date"))
        delivery_date = self._parse_date(s.get("delivery_date"))
        is_delivered = s.get("is_delivered", False)

        dest_parts = [
            s.get("destination_city", ""),
            s.get("destination_state", ""),
            s.get("destination_country", ""),
        ]
        destination = ", ".join(p for p in dest_parts if p)

        days_after_ship = self._calc_days_text(ship_date, delivery_date, now, is_delivered)
        working_days = self._calc_working_days_text(ship_date, delivery_date, now, is_delivered)
        days_after_label = self._calc_days_text(label_date, delivery_date, now, is_delivered)

        historial = self._format_scan_events(s.get("scan_events"))
        recomendacion = self._generate_recommendation(
            sonia_status, ship_date, delivery_date, now, is_delivered
        )

        values = [
            tenant_name,
            s.get("tracking_number", ""),
            str(sonia_status).replace("_", " ").title() if sonia_status else "",
            fedex_status or "",
            str(label_date) if label_date else "",
            str(ship_date) if ship_date else "",
            days_after_ship,
            working_days,
            days_after_label,
            destination,
            historial,
            recomendacion,
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col >= 11))

        if is_delivered:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = DELIVERED_FILL
        elif sonia_status in ("exception", "returned", "customs_hold"):
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = ALERT_FILL

    def _auto_filter(self, ws, last_row: int):
        """Add auto-filter to header row."""
        if last_row >= 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    @staticmethod
    def _parse_date(val) -> Optional[date]:
        """Parse a date from various formats."""
        if val is None:
            return None
        if isinstance(val, date):
            return val
        if isinstance(val, datetime):
            return val.date()
        try:
            return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _business_days(start: date, end: date) -> int:
        """Count business days between two dates."""
        if not start or not end:
            return 0
        days = 0
        current = start
        delta = timedelta(days=1)
        while current < end:
            current += delta
            if current.weekday() < 5:
                days += 1
        return days

    def _calc_days_text(
        self, start: Optional[date], delivery: Optional[date],
        today: date, is_delivered: bool
    ) -> str:
        """Calculate days text like 'ENTREGADO EN 5 DIAS' or '12 DIAS EN TRANSITO'."""
        if not start:
            return ""
        if is_delivered and delivery:
            days = (delivery - start).days
            return f"ENTREGADO EN {days} DIAS"
        else:
            days = (today - start).days
            if days < 0:
                return "PENDIENTE"
            return f"{days} DIAS EN TRANSITO"

    def _calc_working_days_text(
        self, start: Optional[date], delivery: Optional[date],
        today: date, is_delivered: bool
    ) -> str:
        """Calculate working days text."""
        if not start:
            return ""
        if is_delivered and delivery:
            bdays = self._business_days(start, delivery)
            return f"ENTREGADO EN {bdays} DIAS HABILES"
        else:
            bdays = self._business_days(start, today)
            return f"{bdays} DIAS HABILES EN TRANSITO"

    @staticmethod
    def _format_scan_events(scan_events) -> str:
        """Format scan_events JSONB into readable history string."""
        if not scan_events:
            return ""
        if isinstance(scan_events, str):
            import json
            try:
                scan_events = json.loads(scan_events)
            except (json.JSONDecodeError, TypeError):
                return str(scan_events)

        if not isinstance(scan_events, list):
            return str(scan_events)

        parts = []
        for evt in scan_events:
            if isinstance(evt, dict):
                dt = evt.get("date", evt.get("timestamp", ""))
                desc = evt.get("description", evt.get("eventDescription", ""))
                loc = ""
                if evt.get("city"):
                    loc = evt["city"]
                elif evt.get("location"):
                    if isinstance(evt["location"], dict):
                        loc = evt["location"].get("city", "")
                    else:
                        loc = str(evt["location"])
                entry = f"{dt}: {desc}"
                if loc:
                    entry += f" ({loc})"
                parts.append(entry)
            else:
                parts.append(str(evt))

        return " -> ".join(parts)

    @staticmethod
    def _generate_recommendation(
        status: str, ship_date: Optional[date], delivery_date: Optional[date],
        today: date, is_delivered: bool
    ) -> str:
        """Generate SonIA recommendation based on status and timing."""
        if is_delivered and delivery_date and ship_date:
            days = (delivery_date - ship_date).days
            if days <= 7:
                return "Buen tiempo de entrega dentro de lo esperado."
            elif days <= 14:
                return "Entrega dentro de rango aceptable."
            else:
                return f"Entrega demorada ({days} dias). Revisar con FedEx si hay patron."

        if not ship_date:
            return "Sin fecha de envio registrada."

        transit_days = (today - ship_date).days

        if status in ("exception", "delivery_exception"):
            return "ATENCION: Excepcion de entrega. Contactar FedEx inmediatamente."
        elif status in ("returned", "returned_to_sender"):
            return "CRITICO: Paquete devuelto a origen. Accion inmediata requerida."
        elif status in ("customs_hold", "customs"):
            return "En retencion de aduana. Monitorear y preparar documentacion."
        elif transit_days > 21:
            return f"ALERTA: {transit_days} dias en transito. Posible perdida. Abrir reclamo."
        elif transit_days > 14:
            return f"Atencion: {transit_days} dias en transito. Monitorear de cerca."
        elif transit_days > 7:
            return "En transito, dentro de rango normal para envio internacional."
        else:
            return "En transito, tiempo normal."
